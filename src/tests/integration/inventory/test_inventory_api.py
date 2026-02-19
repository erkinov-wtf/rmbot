from datetime import timedelta
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from core.utils.constants import RoleSlug
from inventory.models import (
    Inventory,
    InventoryItem,
    InventoryItemCategory,
    InventoryItemPart,
)
from ticket.models import Ticket

pytestmark = pytest.mark.django_db


LIST_CREATE_URL = "/api/v1/inventory/items/"
CATEGORY_ALL_URL = "/api/v1/inventory/categories/all/"
EXPORT_URL = "/api/v1/inventory/export/"
IMPORT_URL = "/api/v1/inventory/import/"


def detail_url(inventory_item_id: int) -> str:
    return f"/api/v1/inventory/items/{inventory_item_id}/"


def category_detail_url(category_id: int) -> str:
    return f"/api/v1/inventory/categories/{category_id}/"


def _build_import_workbook(*, categories_rows, item_rows) -> bytes:
    workbook = Workbook()
    categories_sheet = workbook.active
    categories_sheet.title = "Categories"
    categories_sheet.append(
        [
            "category_name",
            "part_name",
        ]
    )
    for row in categories_rows:
        categories_sheet.append(
            [
                row.get("category_name"),
                row.get("part_name"),
            ]
        )

    items_sheet = workbook.create_sheet("Inventory Items")
    items_sheet.append(
        [
            "serial_number",
            "name",
            "inventory_name",
            "category_name",
            "status",
            "is_active",
            "category_parts",
        ]
    )
    for row in item_rows:
        items_sheet.append(
            [
                row.get("serial_number"),
                row.get("name"),
                row.get("inventory_name"),
                row.get("category_name"),
                row.get("status"),
                row.get("is_active"),
                row.get("category_parts"),
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_category_all_requires_auth(api_client):
    resp = api_client.get(CATEGORY_ALL_URL)
    assert resp.status_code == 401


def test_category_all_returns_full_category_list(
    authed_client_factory,
    user_factory,
):
    user = user_factory(
        username="inventory_category_reader",
        first_name="Category Reader",
    )
    InventoryItemCategory.objects.create(name="Alpha")
    InventoryItemCategory.objects.create(name="Beta")

    client = authed_client_factory(user)
    resp = client.get(CATEGORY_ALL_URL)

    assert resp.status_code == 200
    names = [entry["name"] for entry in resp.data["data"]]
    assert names[:2] == ["Alpha", "Beta"]


def test_delete_category_returns_error_when_items_exist(
    authed_client_factory,
    user_factory,
    assign_roles,
    inventory_item_factory,
):
    manager_user = user_factory(
        username="inventory_category_delete_blocked",
        first_name="Ops",
    )
    assign_roles(manager_user, RoleSlug.OPS_MANAGER)

    source_category = InventoryItemCategory.objects.create(name="Source Category")
    item = inventory_item_factory(
        serial_number="RM-CAT-DEL-0001",
        category=source_category,
    )
    part = InventoryItemPart.objects.create(
        category=source_category,
        inventory_item=None,
        name="Source Part",
    )

    client = authed_client_factory(manager_user)
    resp = client.delete(category_detail_url(source_category.id))

    assert resp.status_code == 400

    source_category.refresh_from_db()
    item.refresh_from_db()
    part.refresh_from_db()

    assert source_category.deleted_at is None
    assert item.category_id == source_category.id
    assert part.deleted_at is None


def test_delete_category_without_items_soft_deletes_category_and_parts(
    authed_client_factory,
    user_factory,
    assign_roles,
):
    manager_user = user_factory(
        username="inventory_category_delete_allowed",
        first_name="Ops",
    )
    assign_roles(manager_user, RoleSlug.OPS_MANAGER)

    source_category = InventoryItemCategory.objects.create(name="Source No Items")
    part = InventoryItemPart.objects.create(
        category=source_category,
        inventory_item=None,
        name="Part To Archive",
    )

    client = authed_client_factory(manager_user)
    resp = client.delete(category_detail_url(source_category.id))

    assert resp.status_code == 204

    source_category.refresh_from_db()
    part.refresh_from_db()

    assert source_category.deleted_at is not None
    assert part.deleted_at is not None


def test_export_requires_auth(api_client):
    resp = api_client.get(EXPORT_URL)
    assert resp.status_code == 401


def test_export_returns_two_sheet_inventory_workbook(
    authed_client_factory,
    user_factory,
    inventory_item_factory,
):
    user = user_factory(
        username="inventory_export_reader",
        first_name="Export Reader",
    )
    category = InventoryItemCategory.objects.create(name="Export Category")
    item = inventory_item_factory(
        serial_number="RM-EXP-0001",
        name="Export Item",
        category=category,
    )
    InventoryItemPart.objects.create(
        category=category,
        inventory_item=None,
        name="Main Belt",
    )
    client = authed_client_factory(user)

    resp = client.get(EXPORT_URL)

    assert resp.status_code == 200
    assert (
        resp["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in resp["Content-Disposition"]

    workbook = load_workbook(BytesIO(resp.content))
    assert workbook.sheetnames == ["Categories", "Inventory Items"]

    categories_sheet = workbook["Categories"]
    category_headers = [cell.value for cell in categories_sheet[1]]
    assert category_headers == [
        "category_name",
        "part_name",
    ]
    categories_rows = list(categories_sheet.iter_rows(min_row=2, values_only=True))
    assert any(
        row[0] == "Export Category" and row[1] == "Main Belt" for row in categories_rows
    )

    items_sheet = workbook["Inventory Items"]
    item_headers = [cell.value for cell in items_sheet[1]]
    assert item_headers == [
        "serial_number",
        "name",
        "inventory_name",
        "category_name",
        "status",
        "is_active",
        "category_parts",
    ]
    item_rows = list(items_sheet.iter_rows(min_row=2, values_only=True))
    assert any(
        row[0] == item.serial_number
        and row[1] == item.name
        and row[3] == "Export Category"
        and "Main Belt" in str(row[6] or "")
        for row in item_rows
    )


def test_export_accepts_xlsx_accept_header(
    authed_client_factory,
    user_factory,
):
    user = user_factory(
        username="inventory_export_accept_header",
        first_name="Export Accept",
    )
    client = authed_client_factory(user)

    resp = client.get(
        EXPORT_URL,
        HTTP_ACCEPT="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    assert resp.status_code == 200


def test_import_requires_auth(api_client):
    workbook_bytes = _build_import_workbook(
        categories_rows=[{"category_name": "Auth Category", "part_name": "Auth Part"}],
        item_rows=[
            {
                "serial_number": "RM-IMP-AUTH-1",
                "name": "Auth Item",
                "inventory_name": "Auth Inventory",
                "category_name": "Auth Category",
                "status": "ready",
                "is_active": True,
            }
        ],
    )

    upload = SimpleUploadedFile(
        name="inventory_import.xlsx",
        content=workbook_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp = api_client.post(IMPORT_URL, {"file": upload}, format="multipart")
    assert resp.status_code == 401


def test_import_requires_inventory_manager_role(
    authed_client_factory,
    user_factory,
):
    user = user_factory(
        username="inventory_import_regular",
        first_name="Regular",
    )
    workbook_bytes = _build_import_workbook(
        categories_rows=[
            {"category_name": "Import Role Category", "part_name": "Part A"}
        ],
        item_rows=[
            {
                "serial_number": "RM-IMP-ROLE-1",
                "name": "Role Item",
                "inventory_name": "Role Inventory",
                "category_name": "Import Role Category",
                "status": "ready",
                "is_active": True,
            }
        ],
    )
    upload = SimpleUploadedFile(
        name="inventory_import.xlsx",
        content=workbook_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    client = authed_client_factory(user)

    resp = client.post(IMPORT_URL, {"file": upload}, format="multipart")

    assert resp.status_code == 403


def test_import_upserts_categories_parts_and_items(
    authed_client_factory,
    user_factory,
    assign_roles,
    inventory_item_factory,
):
    manager_user = user_factory(
        username="inventory_import_manager",
        first_name="Ops",
    )
    assign_roles(manager_user, RoleSlug.OPS_MANAGER)

    existing_category = InventoryItemCategory.objects.create(name="Phones")
    existing_item = inventory_item_factory(
        serial_number="RM-IMP-0001",
        name="Old Name",
        category=existing_category,
        status="blocked",
        is_active=False,
    )

    workbook_bytes = _build_import_workbook(
        categories_rows=[
            {"category_name": "Phones", "part_name": "Battery"},
            {"category_name": "Laptops", "part_name": "Keyboard"},
        ],
        item_rows=[
            {
                "serial_number": "RM-IMP-0001",
                "name": "Updated Name",
                "inventory_name": "Default Inventory",
                "category_name": "Phones",
                "status": "ready",
                "is_active": True,
            },
            {
                "serial_number": "RM-IMP-0002",
                "name": "New Imported Item",
                "inventory_name": "Imported Inventory",
                "category_name": "Laptops",
                "status": "rented",
                "is_active": False,
            },
        ],
    )
    upload = SimpleUploadedFile(
        name="inventory_import.xlsx",
        content=workbook_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    client = authed_client_factory(manager_user)

    resp = client.post(IMPORT_URL, {"file": upload}, format="multipart")

    assert resp.status_code == 200
    summary = resp.data["data"]
    assert summary["categories_created"] == 1
    assert summary["parts_created"] == 2
    assert summary["items_created"] == 1
    assert summary["items_updated"] == 1

    existing_item.refresh_from_db()
    assert existing_item.name == "Updated Name"
    assert existing_item.status == "ready"
    assert existing_item.is_active is True

    imported_item = InventoryItem.objects.get(serial_number="RM-IMP-0002")
    assert imported_item.name == "New Imported Item"
    assert imported_item.status == "rented"
    assert imported_item.is_active is False
    assert imported_item.category.name == "Laptops"
    assert imported_item.inventory.name == "Imported Inventory"

    assert InventoryItemPart.objects.filter(
        category__name="Phones", name="Battery"
    ).exists()
    assert InventoryItemPart.objects.filter(
        category__name="Laptops", name="Keyboard"
    ).exists()


def test_import_is_atomic_when_workbook_contains_invalid_values(
    authed_client_factory,
    user_factory,
    assign_roles,
):
    manager_user = user_factory(
        username="inventory_import_atomic",
        first_name="Ops",
    )
    assign_roles(manager_user, RoleSlug.OPS_MANAGER)

    workbook_bytes = _build_import_workbook(
        categories_rows=[
            {"category_name": "Atomic Category", "part_name": "Atomic Part"}
        ],
        item_rows=[
            {
                "serial_number": "RM-IMP-BAD-0001",
                "name": "Bad Status Item",
                "inventory_name": "Atomic Inventory",
                "category_name": "Atomic Category",
                "status": "not_a_status",
                "is_active": True,
            }
        ],
    )
    upload = SimpleUploadedFile(
        name="inventory_import.xlsx",
        content=workbook_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    client = authed_client_factory(manager_user)

    resp = client.post(IMPORT_URL, {"file": upload}, format="multipart")

    assert resp.status_code == 400
    assert (
        InventoryItemCategory.objects.filter(name="Atomic Category").exists() is False
    )
    assert InventoryItemPart.objects.filter(name="Atomic Part").exists() is False
    assert (
        InventoryItem.objects.filter(serial_number="RM-IMP-BAD-0001").exists() is False
    )


def test_list_requires_auth(api_client):
    resp = api_client.get(LIST_CREATE_URL)
    assert resp.status_code == 401


def test_create_requires_inventory_item_manager_role(
    authed_client_factory, user_factory
):
    regular_user = user_factory(
        username="regular_inventory_item",
        first_name="Regular",
    )
    client = authed_client_factory(regular_user)

    resp = client.post(LIST_CREATE_URL, {"serial_number": "RM-0200"}, format="json")
    assert resp.status_code == 403


def test_superuser_can_create_inventory_item_without_assigned_roles(
    authed_client_factory, user_factory
):
    superuser = user_factory(
        username="super_inventory_item",
        first_name="Super",
        is_superuser=True,
        is_staff=True,
    )
    client = authed_client_factory(superuser)
    category = InventoryItemCategory.objects.create(name="Super Category")

    resp = client.post(
        LIST_CREATE_URL,
        {"serial_number": "RM-0202", "category": category.id},
        format="json",
    )

    assert resp.status_code == 201
    assert resp.data["data"]["serial_number"] == "RM-0202"


def test_ops_manager_can_create_inventory_item(
    authed_client_factory, user_factory, assign_roles
):
    manager_user = user_factory(
        username="ops_inventory_item",
        first_name="Ops",
    )
    assign_roles(manager_user, RoleSlug.OPS_MANAGER)
    client = authed_client_factory(manager_user)
    category = InventoryItemCategory.objects.create(name="Ops Category")

    resp = client.post(
        LIST_CREATE_URL,
        {"serial_number": "RM-0200", "category": category.id},
        format="json",
    )

    assert resp.status_code == 201
    assert resp.data["data"]["serial_number"] == "RM-0200"
    assert InventoryItem.objects.count() == 1


def test_create_accepts_non_pattern_serial_number(
    authed_client_factory, user_factory, assign_roles
):
    manager_user = user_factory(
        username="ops_inventory_item_regex",
        first_name="Ops",
    )
    assign_roles(manager_user, RoleSlug.OPS_MANAGER)
    client = authed_client_factory(manager_user)
    category = InventoryItemCategory.objects.create(name="Regex Category")

    resp = client.post(
        LIST_CREATE_URL,
        {"serial_number": "inventory_item-0200", "category": category.id},
        format="json",
    )

    assert resp.status_code == 201
    assert resp.data["data"]["serial_number"] == "INVENTORY_ITEM-0200"


def test_create_item_restores_soft_deleted_default_inventory(
    authed_client_factory, user_factory, assign_roles
):
    manager_user = user_factory(
        username="ops_inventory_item_default_inventory_restore",
        first_name="Ops",
    )
    assign_roles(manager_user, RoleSlug.OPS_MANAGER)
    default_inventory = Inventory.objects.create(name="Default Inventory")
    default_inventory.delete()
    category = InventoryItemCategory.objects.create(name="Restore Inventory Category")
    client = authed_client_factory(manager_user)

    resp = client.post(
        LIST_CREATE_URL,
        {"serial_number": "RM-0203", "category": category.id},
        format="json",
    )

    assert resp.status_code == 201
    created_item = InventoryItem.objects.get(pk=resp.data["data"]["id"])
    assert created_item.inventory_id == default_inventory.id
    assert Inventory.objects.filter(pk=default_inventory.pk).exists()
    assert Inventory.all_objects.filter(name="Default Inventory").count() == 1


def test_create_requires_category_and_does_not_create_uncategorized(
    authed_client_factory, user_factory, assign_roles
):
    manager_user = user_factory(
        username="ops_inventory_item_category_required",
        first_name="Ops",
    )
    assign_roles(manager_user, RoleSlug.OPS_MANAGER)
    InventoryItemCategory.all_objects.filter(name="Uncategorized").delete()
    client = authed_client_factory(manager_user)

    resp = client.post(LIST_CREATE_URL, {"serial_number": "RM-0204"}, format="json")

    assert resp.status_code == 400
    assert "category" in resp.data["message"].lower()
    assert (
        InventoryItemCategory.all_objects.filter(name="Uncategorized").exists() is False
    )


def test_list_supports_query_filter_for_serial_number_lookup(
    authed_client_factory, user_factory, inventory_item_factory
):
    user = user_factory(
        username="inventory_item_reader",
        first_name="Reader",
    )
    inventory_item_factory(serial_number="RM-0100")
    inventory_item_factory(serial_number="RM-0101")
    inventory_item_factory(serial_number="RM-0200")
    client = authed_client_factory(user)

    resp = client.get(LIST_CREATE_URL, {"q": "rm-01"})

    assert resp.status_code == 200
    codes = [item["serial_number"] for item in resp.data["results"]]
    assert "RM-0100" in codes
    assert "RM-0101" in codes


def test_list_rejects_short_query(authed_client_factory, user_factory):
    user = user_factory(
        username="inventory_item_reader_short",
        first_name="Reader",
    )
    client = authed_client_factory(user)

    resp = client.get(LIST_CREATE_URL, {"q": "R"})

    assert resp.status_code == 400
    assert "at least 2" in resp.data["message"].lower()


def test_list_supports_extended_filters(
    authed_client_factory,
    user_factory,
    inventory_item_factory,
):
    user = user_factory(
        username="inventory_item_reader_filters",
        first_name="Reader",
    )
    active_inventory_item = inventory_item_factory(
        serial_number="RM-0300", status="ready", is_active=True
    )
    inactive_inventory_item = inventory_item_factory(
        serial_number="RM-0301", status="blocked", is_active=False
    )
    master = user_factory(
        username="inventory_item_filter_master",
        first_name="Master",
    )
    Ticket.objects.create(
        inventory_item=active_inventory_item,
        master=master,
        status="new",
        title="Filter test ticket",
    )
    client = authed_client_factory(user)

    old_date = timezone.now() - timedelta(days=10)
    InventoryItem.all_objects.filter(pk=inactive_inventory_item.pk).update(
        created_at=old_date
    )

    status_filtered = client.get(
        LIST_CREATE_URL, {"status": "ready", "is_active": "true"}
    )
    assert status_filtered.status_code == 200
    status_codes = [item["serial_number"] for item in status_filtered.data["results"]]
    assert status_codes == ["RM-0300"]

    active_ticket_filtered = client.get(LIST_CREATE_URL, {"has_active_ticket": "true"})
    assert active_ticket_filtered.status_code == 200
    active_ticket_codes = [
        item["serial_number"] for item in active_ticket_filtered.data["results"]
    ]
    assert active_ticket_codes == ["RM-0300"]

    created_range_filtered = client.get(
        LIST_CREATE_URL,
        {
            "created_to": (timezone.now().date() - timedelta(days=5)).isoformat(),
        },
    )
    assert created_range_filtered.status_code == 200
    created_range_codes = [
        item["serial_number"] for item in created_range_filtered.data["results"]
    ]
    assert created_range_codes == ["RM-0301"]


def test_update_requires_inventory_item_manager_role(
    authed_client_factory, user_factory, inventory_item_factory
):
    regular_user = user_factory(
        username="inventory_item_update_regular",
        first_name="Regular",
    )
    inventory_item = inventory_item_factory(serial_number="RM-0400")
    client = authed_client_factory(regular_user)

    resp = client.patch(
        detail_url(inventory_item.id), {"status": "blocked"}, format="json"
    )

    assert resp.status_code == 403


def test_ops_manager_can_update_inventory_item(
    authed_client_factory, user_factory, assign_roles, inventory_item_factory
):
    manager_user = user_factory(
        username="inventory_item_update_ops",
        first_name="Ops",
    )
    assign_roles(manager_user, RoleSlug.OPS_MANAGER)
    inventory_item = inventory_item_factory(
        serial_number="RM-0401", status="ready", is_active=True
    )
    client = authed_client_factory(manager_user)

    resp = client.patch(
        detail_url(inventory_item.id),
        {"status": "blocked", "is_active": False},
        format="json",
    )

    assert resp.status_code == 200
    assert resp.data["data"]["status"] == "blocked"
    assert resp.data["data"]["is_active"] is False

    inventory_item.refresh_from_db()
    assert inventory_item.status == "blocked"
    assert inventory_item.is_active is False


def test_delete_requires_inventory_item_manager_role(
    authed_client_factory, user_factory, inventory_item_factory
):
    regular_user = user_factory(
        username="inventory_item_delete_regular",
        first_name="Regular",
    )
    inventory_item = inventory_item_factory(serial_number="RM-0500")
    client = authed_client_factory(regular_user)

    resp = client.delete(detail_url(inventory_item.id))

    assert resp.status_code == 403


def test_ops_manager_can_soft_delete_inventory_item(
    authed_client_factory, user_factory, assign_roles, inventory_item_factory
):
    manager_user = user_factory(
        username="inventory_item_delete_ops",
        first_name="Ops",
    )
    assign_roles(manager_user, RoleSlug.OPS_MANAGER)
    inventory_item = inventory_item_factory(serial_number="RM-0501")
    client = authed_client_factory(manager_user)

    resp = client.delete(detail_url(inventory_item.id))

    assert resp.status_code == 204
    assert resp.content == b""
    assert InventoryItem.objects.filter(pk=inventory_item.pk).exists() is False
    assert InventoryItem.all_objects.filter(
        pk=inventory_item.pk, deleted_at__isnull=False
    ).exists()
