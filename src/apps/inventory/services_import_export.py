from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from django.db import transaction

from core.utils.constants import InventoryItemStatus
from inventory.models import (
    InventoryItem,
    InventoryItemCategory,
    InventoryItemPart,
)
from inventory.services import InventoryItemService

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass
class InventoryImportSummary:
    categories_created: int = 0
    categories_updated: int = 0
    parts_created: int = 0
    parts_updated: int = 0
    items_created: int = 0
    items_updated: int = 0

    def as_dict(self) -> dict[str, int]:
        return {
            "categories_created": self.categories_created,
            "categories_updated": self.categories_updated,
            "parts_created": self.parts_created,
            "parts_updated": self.parts_updated,
            "items_created": self.items_created,
            "items_updated": self.items_updated,
        }


class InventoryImportExportService:
    CATEGORIES_SHEET = "Categories"
    ITEMS_SHEET = "Inventory Items"

    CATEGORY_HEADERS = (
        "category_name",
        "part_name",
    )
    ITEM_HEADERS = (
        "serial_number",
        "name",
        "category_name",
        "status",
        "is_active",
    )

    VALID_STATUSES = {choice for choice, _ in InventoryItemStatus.choices}
    TRUTHY = {"1", "true", "yes", "y", "on"}
    FALSY = {"0", "false", "no", "n", "off"}

    @classmethod
    def export_workbook_bytes(cls) -> bytes:
        from openpyxl import Workbook

        workbook = Workbook()
        categories_sheet = workbook.active
        categories_sheet.title = cls.CATEGORIES_SHEET
        categories_sheet.append(list(cls.CATEGORY_HEADERS))

        categories = list(
            InventoryItemCategory.domain.get_queryset().order_by("name", "id")
        )
        parts = list(
            InventoryItemPart.domain.get_queryset()
            .select_related("category")
            .order_by("category__name", "name", "id")
        )
        parts_by_category: dict[int, list[InventoryItemPart]] = {}
        for part in parts:
            parts_by_category.setdefault(part.category_id, []).append(part)

        for category in categories:
            category_parts = parts_by_category.get(category.id, [])
            if not category_parts:
                categories_sheet.append(
                    [
                        category.name,
                        "",
                    ]
                )
                continue

            for part in category_parts:
                categories_sheet.append(
                    [
                        category.name,
                        part.name,
                    ]
                )

        items_sheet = workbook.create_sheet(title=cls.ITEMS_SHEET)
        items_sheet.append(list(cls.ITEM_HEADERS))
        items = list(
            InventoryItem.domain.get_queryset()
            .select_related("category")
            .order_by("serial_number", "id")
        )
        for item in items:
            items_sheet.append(
                [
                    item.serial_number,
                    item.name,
                    item.category.name,
                    item.status,
                    bool(item.is_active),
                ]
            )

        cls._autosize_columns(categories_sheet)
        cls._autosize_columns(items_sheet)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @classmethod
    def import_workbook_bytes(cls, *, workbook_bytes: bytes) -> dict[str, int]:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
        if cls.CATEGORIES_SHEET not in workbook.sheetnames:
            raise ValueError(f"Missing required sheet: {cls.CATEGORIES_SHEET}")
        if cls.ITEMS_SHEET not in workbook.sheetnames:
            raise ValueError(f"Missing required sheet: {cls.ITEMS_SHEET}")

        categories_rows = cls._read_sheet_rows(
            worksheet=workbook[cls.CATEGORIES_SHEET],
            expected_headers=cls.CATEGORY_HEADERS,
        )
        items_rows = cls._read_sheet_rows(
            worksheet=workbook[cls.ITEMS_SHEET],
            expected_headers=cls.ITEM_HEADERS,
        )

        summary = InventoryImportSummary()

        with transaction.atomic():
            category_cache: dict[str, InventoryItemCategory] = {}

            for row_number, row in categories_rows:
                category_name = cls._required_string(
                    value=row.get("category_name"),
                    field_name="category_name",
                    row_number=row_number,
                )
                category, category_created = cls._upsert_category(name=category_name)
                if category_created:
                    summary.categories_created += 1
                else:
                    summary.categories_updated += 1
                category_cache[category.name.casefold()] = category

                part_name = cls._optional_string(row.get("part_name"))
                if not part_name:
                    continue

                _, part_created = cls._upsert_part(
                    category=category, part_name=part_name
                )
                if part_created:
                    summary.parts_created += 1
                else:
                    summary.parts_updated += 1

            for row_number, row in items_rows:
                serial_number = cls._required_string(
                    value=row.get("serial_number"),
                    field_name="serial_number",
                    row_number=row_number,
                )
                normalized_serial = InventoryItemService.normalize_serial_number(
                    serial_number
                )
                if not normalized_serial:
                    raise ValueError(
                        f"Row {row_number}: serial_number must not be empty."
                    )

                category_name = cls._required_string(
                    value=row.get("category_name"),
                    field_name="category_name",
                    row_number=row_number,
                )
                category_key = category_name.casefold()
                category = category_cache.get(category_key)
                if category is None:
                    category, category_created = cls._upsert_category(
                        name=category_name
                    )
                    if category_created:
                        summary.categories_created += 1
                    else:
                        summary.categories_updated += 1
                    category_cache[category_key] = category

                inventory = InventoryItemService.get_default_inventory()

                status = cls._parse_status(
                    value=row.get("status"),
                    row_number=row_number,
                )
                is_active = cls._parse_bool(
                    value=row.get("is_active"),
                    row_number=row_number,
                    field_name="is_active",
                    default_value=True,
                )
                name = cls._optional_string(row.get("name")) or normalized_serial

                item, item_created = cls._upsert_item(
                    serial_number=normalized_serial,
                    name=name,
                    inventory=inventory,
                    category=category,
                    status=status,
                    is_active=is_active,
                )
                if item_created:
                    summary.items_created += 1
                else:
                    summary.items_updated += 1

        return summary.as_dict()

    @classmethod
    def _read_sheet_rows(
        cls,
        *,
        worksheet,
        expected_headers: tuple[str, ...],
    ) -> list[tuple[int, dict[str, Any]]]:
        raw_headers = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
        )
        headers = [cls._optional_string(value) or "" for value in raw_headers]
        missing_headers = [
            header for header in expected_headers if header not in headers
        ]
        if missing_headers:
            raise ValueError(
                f"Sheet '{worksheet.title}' is missing headers: {', '.join(missing_headers)}"
            )

        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if row_values is None:
                continue
            if all(cls._optional_string(value) in (None, "") for value in row_values):
                continue

            mapped_row = {
                headers[index]: row_values[index] if index < len(row_values) else None
                for index in range(len(headers))
            }
            rows.append((row_number, mapped_row))
        return rows

    @classmethod
    def _upsert_category(cls, *, name: str) -> tuple[InventoryItemCategory, bool]:
        existing = (
            InventoryItemCategory.all_objects.filter(name__iexact=name)
            .order_by("id")
            .first()
        )
        if existing is None:
            return InventoryItemCategory.objects.create(name=name), True

        update_fields: list[str] = []
        if existing.deleted_at is not None:
            existing.deleted_at = None
            update_fields.append("deleted_at")
        if existing.name != name:
            existing.name = name
            update_fields.append("name")
        if update_fields:
            existing.save(update_fields=update_fields)
        return existing, False

    @classmethod
    def _upsert_part(
        cls, *, category: InventoryItemCategory, part_name: str
    ) -> tuple[InventoryItemPart, bool]:
        existing = (
            InventoryItemPart.all_objects.filter(
                category_id=category.id,
                name__iexact=part_name,
            )
            .order_by("id")
            .first()
        )
        if existing is None:
            return (
                InventoryItemPart.objects.create(
                    category=category,
                    inventory_item=None,
                    name=part_name,
                ),
                True,
            )

        update_fields: list[str] = []
        if existing.deleted_at is not None:
            existing.deleted_at = None
            update_fields.append("deleted_at")
        if existing.name != part_name:
            existing.name = part_name
            update_fields.append("name")
        if existing.category_id != category.id:
            existing.category = category
            update_fields.append("category")
        if existing.inventory_item_id is not None:
            existing.inventory_item = None
            update_fields.append("inventory_item")
        if update_fields:
            existing.save(update_fields=update_fields)
        return existing, False

    @classmethod
    def _upsert_item(
        cls,
        *,
        serial_number: str,
        name: str,
        inventory: Inventory,
        category: InventoryItemCategory,
        status: str,
        is_active: bool,
    ) -> tuple[InventoryItem, bool]:
        existing = (
            InventoryItem.all_objects.filter(serial_number__iexact=serial_number)
            .order_by("id")
            .first()
        )
        if existing is None:
            return (
                InventoryItem.objects.create(
                    serial_number=serial_number,
                    name=name,
                    inventory=inventory,
                    category=category,
                    status=status,
                    is_active=is_active,
                ),
                True,
            )

        update_fields: list[str] = []
        if existing.deleted_at is not None:
            existing.deleted_at = None
            update_fields.append("deleted_at")
        if existing.serial_number != serial_number:
            existing.serial_number = serial_number
            update_fields.append("serial_number")
        if existing.name != name:
            existing.name = name
            update_fields.append("name")
        if existing.inventory_id != inventory.id:
            existing.inventory = inventory
            update_fields.append("inventory")
        if existing.category_id != category.id:
            existing.category = category
            update_fields.append("category")
        if existing.status != status:
            existing.status = status
            update_fields.append("status")
        if existing.is_active is not is_active:
            existing.is_active = is_active
            update_fields.append("is_active")
        if update_fields:
            existing.save(update_fields=update_fields)
        return existing, False

    @classmethod
    def _parse_status(cls, *, value: Any, row_number: int) -> str:
        status = cls._optional_string(value)
        if not status:
            return InventoryItemStatus.READY
        normalized = status.strip().lower()
        if normalized not in cls.VALID_STATUSES:
            raise ValueError(
                f"Row {row_number}: unsupported status '{status}'. "
                f"Allowed values: {', '.join(sorted(cls.VALID_STATUSES))}."
            )
        return normalized

    @classmethod
    def _parse_bool(
        cls,
        *,
        value: Any,
        row_number: int,
        field_name: str,
        default_value: bool,
    ) -> bool:
        if value is None:
            return default_value
        if isinstance(value, bool):
            return value
        parsed_text = cls._optional_string(value)
        if parsed_text is None or parsed_text == "":
            return default_value

        normalized = parsed_text.strip().casefold()
        if normalized in cls.TRUTHY:
            return True
        if normalized in cls.FALSY:
            return False
        raise ValueError(
            f"Row {row_number}: {field_name} must be boolean-like (true/false)."
        )

    @staticmethod
    def _required_string(*, value: Any, field_name: str, row_number: int) -> str:
        parsed = InventoryImportExportService._optional_string(value)
        if not parsed:
            raise ValueError(f"Row {row_number}: {field_name} is required.")
        return parsed

    @staticmethod
    def _optional_string(value: Any) -> str | None:
        if value is None:
            return None
        parsed = str(value).strip()
        if not parsed:
            return None
        return parsed

    @staticmethod
    def _autosize_columns(worksheet) -> None:
        for column in worksheet.columns:
            values = [str(cell.value) for cell in column if cell.value is not None]
            max_length = max((len(value) for value in values), default=0)
            column_letter = column[0].column_letter
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12), 60
            )
